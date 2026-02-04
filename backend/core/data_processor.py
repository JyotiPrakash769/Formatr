import os
import json
import csv
import openpyxl
import xlrd

class DataProcessor:
    @staticmethod
    def csv_to_json(input_path: str, output_dir: str) -> str:
        """Convert CSV to JSON without Pandas"""
        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.json")
        
        data = []
        with open(input_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
            
        return output_path
    
    @staticmethod
    def json_to_csv(input_path: str, output_dir: str) -> str:
        """Convert JSON to CSV without Pandas"""
        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.csv")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if not data:
            return output_path
            
        keys = data[0].keys()
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(data)
            
        return output_path
    
    @staticmethod
    def csv_to_excel(input_path: str, output_dir: str) -> str:
        """Convert CSV to Excel (XLSX) without Pandas"""
        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        with open(input_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def excel_to_csv(input_path: str, output_dir: str) -> str:
        """Convert Excel (XLSX/XLS) to CSV without Pandas"""
        filename = os.path.basename(input_path)
        name, ext = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.csv")
        
        data = []
        if ext.lower() == '.xls':
            # Legacy XLS handling
            rb = xlrd.open_workbook(input_path)
            sheet = rb.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                data.append(sheet.row_values(row_idx))
        else:
            # Modern XLSX handling
            wb = openpyxl.load_workbook(input_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)
            
        return output_path
    
    @staticmethod
    def excel_to_json(input_path: str, output_dir: str) -> str:
        """Convert Excel (XLSX/XLS) to JSON without Pandas"""
        filename = os.path.basename(input_path)
        name, ext = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.json")
        
        rows = []
        if ext.lower() == '.xls':
            rb = xlrd.open_workbook(input_path)
            sheet = rb.sheet_by_index(0)
            header = sheet.row_values(0)
            for row_idx in range(1, sheet.nrows):
                row_data = dict(zip(header, sheet.row_values(row_idx)))
                rows.append(row_data)
        else:
            wb = openpyxl.load_workbook(input_path, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter)
            for row in rows_iter:
                row_data = dict(zip(header, row))
                rows.append(row_data)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2)
            
        return output_path
    
    @staticmethod
    def xls_to_xlsx(input_path: str, output_dir: str) -> str:
        """Convert legacy Excel (XLS) to modern Excel (XLSX) without Pandas"""
        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        output_path = os.path.join(output_dir, f"{name}.xlsx")
        
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        
        rb = xlrd.open_workbook(input_path)
        sheet = rb.sheet_by_index(0)
        
        for row_idx in range(sheet.nrows):
            ws_new.append(sheet.row_values(row_idx))
            
        wb_new.save(output_path)
        return output_path
